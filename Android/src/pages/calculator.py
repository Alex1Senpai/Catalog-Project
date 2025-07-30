import flet as ft
from pages.tovari import IMGS_DIR, DEFAULT_IMAGE_PATH
import openpyxl
from datetime import datetime
import os
from pathlib import Path
import re
import platform
import shutil

if platform.system() == "Windows":
    BASE_DIR = Path(__file__).resolve().parent.parent
else:
    BASE_DIR = Path(os.getenv("ANDROID_PRIVATE", ""))
SAVE_DIR = BASE_DIR / "backend"
SHABLON_DIR = SAVE_DIR / "SHABLON"
HISTORY_DIR = SAVE_DIR / "history"
DEFAULT_DB_PATH = SAVE_DIR / "back.db"

EXTERNAL_SELECTED_DIR = "external_selected_dir"

def generate_contract(e: ft.ControlEvent, page: ft.Page):
    full_name = page.session.get("full_name") or "Не указано"

    client_name = ft.TextField(label="Ф.И.О Клиента", width=300)
    client_address = ft.TextField(label="Адрес клиента", width=300)
    client_phone = ft.TextField(label="Телефон", width=300)

    def close_dlg(e):
        page.close(dlg)
        page.update()

    def save_contract(e):
        if not all([client_name.value, client_address.value, client_phone.value]):
            page.open(ft.SnackBar(ft.Text("Заполните все поля!")))
            return

        try:
            if not (SHABLON_DIR / "shablon.xlsx").exists():
                raise FileNotFoundError("Файл шаблона не найден!")

            selected_dir = page.session.get(EXTERNAL_SELECTED_DIR)
            if not selected_dir:
                raise Exception("Сначала выберите папку для сохранения файла")

            wb = openpyxl.load_workbook(SHABLON_DIR / "shablon.xlsx")
            ws = wb.active

            ws['D10'] = client_name.value
            ws['D11'] = client_address.value
            ws['D12'] = client_phone.value
            ws['F5'] = datetime.now().strftime("%d.%m.%Y %H:%M")
            ws['E14'] = full_name

            selected_items = page.session.get("selected_items") or []
            total_amount = sum(item['selling_price'] * item['quantity'] for item in selected_items)
            ws['F8'] = total_amount

            if selected_items:
                start_row = 8
                end_row = 14
                available_rows = end_row - start_row
                if len(selected_items) > available_rows:
                    raise ValueError("Слишком много товаров для шаблона")

                ws.insert_rows(start_row, amount=len(selected_items))

                for idx, item in enumerate(selected_items, start=start_row):
                    ws[f'A{idx}'] = item.get('category_name', '')
                    ws[f'B{idx}'] = item.get('name', '')
                    ws[f'C{idx}'] = item.get('unit', 'шт')
                    ws[f'D{idx}'] = item.get('quantity', 0)
                    ws[f'E{idx}'] = item.get('selling_price', 0)
                    ws[f'F{idx}'] = item['selling_price'] * item['quantity']

            username_raw = page.session.get("username")
            username_str = "unknown"
            if username_raw is not None:
                username_str = str(getattr(username_raw, 'value', username_raw))

            safe_username = re.sub(r'[\\/*?:"<>|]', "", username_str).strip()
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            filename = f"{safe_username}_{timestamp}.xlsx"

            external_path = Path(selected_dir) / filename
            wb.save(external_path)

            page.open(ft.SnackBar(ft.Text(f"Договор сохранен: {filename}")))
        except Exception as ex:
            import traceback
            traceback.print_exc()
            page.open(ft.SnackBar(ft.Text(f"Ошибка: {str(ex)}")))
        finally:
            close_dlg(e)

    dlg = ft.AlertDialog(
        title=ft.Text("Данные клиента"),
        content=ft.Column(
            controls=[client_name, client_address, client_phone],
            tight=True,
        ),
        actions=[
            ft.TextButton("Сохранить", on_click=save_contract),
            ft.TextButton("Отмена", on_click=close_dlg),
        ],
    )

    page.open(dlg)
    page.update()

def build_calculate_content(page, calculate_content_container, show_categories, update_item_list=None):
    items_column = ft.Column(scroll=ft.ScrollMode.ALWAYS, expand=True)
    selected_items = page.session.get("selected_items") or []
    
    # Общая сумма
    total_text = ft.Text(
        value=f"Общая сумма: {sum(item['selling_price'] * item['quantity'] for item in selected_items):,} UZS".replace(",", " "),
        size=16,
        weight=ft.FontWeight.BOLD,
        color=ft.Colors.GREEN_800
    )

    def update_total():
        """Обновляет сумму без перерисовки всего интерфейса"""
        total = sum(item['selling_price'] * item['quantity'] for item in selected_items)
        total_text.value = f"Общая сумма: {total:,} UZS".replace(",", " ")
        total_text.update()

    # Фабрика обработчиков
    def create_handlers(item_id, quantity_control):
        def increment_handler(e):
            for item in selected_items:
                if item['id'] == item_id:
                    item['quantity'] += 1
                    quantity_control.value = str(item['quantity'])
                    quantity_control.update()
                    break
            update_total()

        def decrement_handler(e):
            entry_to_remove = None
            for idx, item in enumerate(selected_items):
                if item['id'] == item_id:
                    item['quantity'] -= 1
                    if item['quantity'] <= 0:
                        entry_to_remove = (idx, item_id)
                    else:
                        quantity_control.value = str(item['quantity'])
                        quantity_control.update()
                    break
            
            if entry_to_remove:
                # Удаляем элемент из данных
                selected_items.pop(entry_to_remove[0])
                
                # Удаляем из UI
                for container in items_column.controls[:]:
                    if container.data == entry_to_remove[1]:
                        items_column.controls.remove(container)
                        break
                
                # Обновляем сессию и интерфейс
                page.session.set("selected_items", selected_items)
                items_column.update()
                if update_item_list:
                    update_item_list()  # Обновляем цвет карточек
                
            update_total()

        return increment_handler, decrement_handler

    # Строим список товаров
    for item in selected_items:
        image_id = item.get('image_id')
        image_path = IMGS_DIR / f"{image_id}.jpg" if image_id else DEFAULT_IMAGE_PATH
        if not image_path.exists():
            image_path = DEFAULT_IMAGE_PATH

        quantity_text = ft.Text(str(item['quantity']), width=40, text_align=ft.TextAlign.CENTER)
        increment, decrement = create_handlers(item['id'], quantity_text)

        container = ft.Container(
            ft.Row(
                controls=[
                    ft.Image(
                        src=str(image_path),
                        width=50,
                        height=50,
                        fit=ft.ImageFit.COVER,
                        border_radius=5
                    ),
                    ft.Column(
                        [
                            ft.Text(item['name'], width=150, max_lines=2),
                            ft.Text(f"{item['selling_price']:,} UZS".replace(",", " "), size=12),
                        ],
                        alignment=ft.MainAxisAlignment.CENTER,
                        expand=True
                    ),
                    ft.Row(
                        [
                            ft.IconButton(
                                ft.Icons.REMOVE_CIRCLE_OUTLINED,
                                on_click=decrement,
                                icon_size=20,
                                tooltip="Уменьшить"
                            ),
                            quantity_text,
                            ft.IconButton(
                                ft.Icons.ADD_CIRCLE_OUTLINED,
                                on_click=increment,
                                icon_size=20,
                                tooltip="Увеличить"
                            ),
                        ],
                        alignment=ft.MainAxisAlignment.END,
                        spacing=0
                    )
                ],
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
                spacing=10
            ),
            padding=5,
            border=ft.border.all(1, ft.Colors.GREY_300),
            border_radius=10,
            data=item['id']  # Сохраняем ID для поиска
        )
        items_column.controls.append(container)

    # Кнопка "Очистить всё"
    def clear_all(e):
        selected_items.clear()
        items_column.controls.clear()
        page.session.set("selected_items", [])
        if update_item_list:
            update_item_list()
        update_total()
        items_column.update()

    clear_button = ft.ElevatedButton(
        "Очистить всё",
        on_click=clear_all,
        icon=ft.Icons.DELETE_FOREVER_ROUNDED,
        bgcolor=ft.Colors.RED_500,
        color=ft.Colors.WHITE,
    )

    # Контейнер с кнопками и суммой
    bottom_buttons = ft.Column(
        controls=[
            ft.Divider(),
            ft.Row(
                [total_text],
                alignment=ft.MainAxisAlignment.END,
                vertical_alignment=ft.CrossAxisAlignment.CENTER
            ),
            ft.Row(
                [
                    clear_button,
                    ft.Container(expand=True),
                    ft.ElevatedButton(
                        "Договор",
                        icon=ft.Icons.ASSIGNMENT,
                        bgcolor=ft.Colors.BLUE_500,
                        color=ft.Colors.WHITE,
                        on_click=lambda e: generate_contract(e, page)
                    )
                ],
                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
            )
        ],
        spacing=10
    )

    return ft.Container(
        content=ft.Column(
            controls=[items_column, bottom_buttons],
            expand=True
        ),
        padding=10
    )